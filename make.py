import csv, sys
import openpyxl as opx

class ReadData:
    """
    data 폴더 내부의 application.csv 파일과, timetable.xlsx 파일을 읽고 각각 list, dict형태로 반환하는 함수를 제공한다.
    """
    def __init__(self):
        pass

    def read_application(self) -> list:
        """
        data 폴더의 application.csv파일에 대해 각 지원자의 입력 정보를 dict로 읽고 모든 지원자의 정보를 가지는 list를 반환한다.
        이때, 지원자의 이름과 전화번호 끝 네자리를 결합한 pk key값을 추가한다.
        
        * error case
        FileNotFoundError -> 관련 텍스트 출력 후 프로그램 종료

        * return example
        [
            {
                '타임스탬프': '2020/03/20 11:58:49 오후 GMT+9',
                '전화 번호': '01012345678',
                '성함': '홍길동',
                '재학중인 학교': '서울과학기술대학교',
                '학과 (ex. ~학과 )': '컴퓨터공학과',
                '지원 동기 (500자 내외)': '...',
                '자신이 만들고 싶은 서비스에 대해 간략하게 알려주세요 (500자 내외)': '...',
                '멋쟁이 사자처럼 활동을 하면서 얻고 싶은 것 (300자 내외)': '...',
                '자신이 다뤄본 컴퓨터 언어': 'C, C++',
                '저희와 함께 하시겠습니까?': '예',
                'pk': '홍길동5678'
            },{
                ...
            }, ...
        ]
        """
        try:
            application_file_name = './data/application.csv'
            application_list = []
            with open(application_file_name, newline='') as application_file:
                application_dict_reader = csv.DictReader(application_file)
                for row in application_dict_reader:
                    row['pk'] = row['성함']+row['전화 번호'][-4:]
                    application_list.append(row)
            return application_list
        except FileNotFoundError:
            print("[ERROR] application.csv 파일이 존재하지 않습니다.")
            sys.exit()
            
    def read_timetable(self) -> dict:
        """
        data 폴더의 timetable.xlsx파일에 대해 오프라인/온라인 별로 읽어서 시간순서대로 지원자의 이름을 list로 가지는 dict를 반환한다.

        * error case
        FileNotFoundError -> 관련 텍스트 출력 후 프로그램 종료
        KeyError -> 관련 텍스트 출력 후 프로그램 종료

        * return example
        {
            'offline':['홍길동','김철수',...],
            'online':[...]
        }
        """
        try:
            timetable_file_name = './data/timetable.xlsx'
            return_dict = {
                'offline':[],
                'online':[]
            }
            timetable_wb = opx.load_workbook(filename = timetable_file_name)
            offline_ws = timetable_wb['오프라인']
            online_ws = timetable_wb['온라인']
            for row_num in range(2,4):
                for col_num in range(2,12):
                    if col_num <=  7:
                        applicant_list = (online_ws.cell(row=row_num, column=col_num).value).split('\n')
                        return_dict['online'] += filter(lambda x: x!='', applicant_list)
                    applicant_list = (offline_ws.cell(row=row_num, column=col_num).value).split('\n')
                    return_dict['offline'] += filter(lambda x: x!='', applicant_list)
            return return_dict
        except FileNotFoundError:
            print("[ERROR] timetable.xlsx 파일이 존재하지 않습니다.")
            sys.exit()
        except KeyError:
            print("[ERROR] sheet 이름이 '오프라인' 또는 '온라인' 으로 되어있는지 확인해주세요.")
            sys.exit()

        pass

class DivideOnOffLineApplicant:
    """
    클래스 생성시 application를 list로, timetable을 dict로 받아서 초기화하며
    각 지원자를 offline과 online으로 나누어 list로 반환하는 함수를 제공한다.
    """

    application: list
    timetable: dict

    def __init__(self, application:list, timetable:dict):
        self.application = application
        self.timetable = timetable

    def divide_offline(self) -> list:
        """
        class 생성시 초기화한 application list와 timetable dict을 이용하여
        offline 대상자의 dict정보를 list화 하여 반환한다.
        """
        return_list = []
        for a in self.application:
            if a['성함'] in self.timetable['offline']:
                return_list.append(a)
        return return_list

    def divide_online(self) -> list:
        """
        class 생성시 초기화한 application list와 timetable dict을 이용하여
        online 대상자의 dict정보를 list화 하여 반환한다.
        """
        return_list = []
        for a in self.application:
            if a['성함'] in self.timetable['online']:
                return_list.append(a)
        return return_list

class CreateData:
    """
    data 폴더에 online_data.xlsx 파일과 offline_data.xlsx 파일을 생성하는 함수를 제공한다.
    기존에 동일한 이름의 파일이 존재할 때에는 덮어쓰기를 한다.
    """

    on_line_applicant_list: list
    off_line_applicant_list: list

    def __init__(self, off_line_applicant_list:list, on_line_applicant_list:list):
        self.off_line_applicant_list = off_line_applicant_list
        self.on_line_applicant_list = on_line_applicant_list

    def create_data(self) -> bool:
        """
        클래스 생성시 초기화한 on_line_applicant_list와 off_line_applicant_list를 통해 각각 online_data.xlsx파일, offline_data.xlsx파일로 생성한다.
        """
        try:
            # offline
            offline_filename = './data/offline_data.xlsx'
            offline_wb = opx.Workbook()
            offline_ws = offline_wb.active
            offline_ws.append(['타임스탬프', '전화 번호', '성함', '재학중인 학교', '학과 (ex. ~학과 )', '지원 동기 (500자 내외)', '자신이 만들고 싶은 서비스에 대해 간략하게 알려주세요 (500자 내외)', '멋쟁이 사자처럼 활동을 하면서 얻고 싶은 것 (300자 내외)', '자신이 다뤄본 컴퓨터 언어', '저희와 함께 하시겠습니까?', 'pk'])
            for applicant in self.off_line_applicant_list:
                offline_ws.append(list(applicant.values()))
            offline_wb.save(filename=offline_filename)

            # online
            online_filename = './data/online_data.xlsx'
            online_wb = opx.Workbook()
            online_ws = online_wb.active
            online_ws.append(['타임스탬프', '전화 번호', '성함', '재학중인 학교', '학과 (ex. ~학과 )', '지원 동기 (500자 내외)', '자신이 만들고 싶은 서비스에 대해 간략하게 알려주세요 (500자 내외)', '멋쟁이 사자처럼 활동을 하면서 얻고 싶은 것 (300자 내외)', '자신이 다뤄본 컴퓨터 언어', '저희와 함께 하시겠습니까?', 'pk'])
            for applicant in self.on_line_applicant_list:
                online_ws.append(list(applicant.values()))
            online_wb.save(filename=online_filename)

            return True
        except:
            return False
        
        


def main() -> None:
    """
    main 함수.
    1. ReadData 클래스를 통해 application.csv 파일과 timetable.xlsx 파일을 읽어서 각각 list, dict 형태로 가져온다.
    2. 이후 timetable.xlsx 파일을 기준으로, application.csv에 있는 지원자의 정보를 오프라인/온라인으로 구분한다
    3. CreateData 클래스를 통해 offline_data.xlsx 파일과 online_data.xlsx 파일을 생성한다.
    """
    # 1. Read
    read_data_cls = ReadData()
    application_list = read_data_cls.read_application()
    timetable_dict = read_data_cls.read_timetable()

    # 2. Divied
    divide_on_off_cls = DivideOnOffLineApplicant(application_list, timetable_dict)
    divided_offline_applicants = divide_on_off_cls.divide_offline()
    divided_online_applicants = divide_on_off_cls.divide_online()

    # 3. Create
    create_data_cls = CreateData(divided_offline_applicants, divided_online_applicants)
    create_result = create_data_cls.create_data()
    if create_result:
        print("data 폴더에 정상적으로 파일을 생성하였습니다.")
    else:
        print("파일 생성에 실패하였습니다.")

if __name__ == '__main__':
    main()